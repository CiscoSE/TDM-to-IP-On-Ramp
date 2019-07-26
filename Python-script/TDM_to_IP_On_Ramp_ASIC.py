# -*- coding: utf-8 -*-

"""
Copyright (c) 2019 Cisco and/or its affiliates.

This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.0 (the "License"). You may obtain a copy of the
License at

               https://developer.cisco.com/docs/licenses

All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.
"""


from __future__ import absolute_import, division, print_function


__author__ = "Tahsin A Chowdhury <tchowdhu@cisco.com>"
__collaborator__= "Alan Liew <aliew@cisco.com>"

__copyright__ = "Copyright (c) 2018 Cisco and/or its affiliates."
__license__ = "Cisco Sample Code License, Version 1.0"


from openpyxl import load_workbook
from texttable import Texttable
import warnings
import telnetlib
import socket
import sys

IOS_XE = 0

warnings.simplefilter('ignore')
wb = load_workbook(filename=sys.argv[1])
warnings.simplefilter('ignore')

sheetname = wb.sheetnames[0]
sheetN = wb[sheetname]

eRow = 1
eColumn = 2


# print(sum([bin(int(x)).count('1') for x in '255.255.255.252'.split('.')]))

# user defined function:
# name: ip_input_format(inIpAddress)
# input: inApAddress: Ip address taken as input
# function: returns Host IP address after checking the correct format
def ip_input_format(inIpAddress):
    while (True):
        ip_address = inIpAddress
        try:
            socket.inet_aton(ip_address)
            # print(ip_address)
            break
        except socket.error:
            print("{}: Invalid IP Input. Update/Correct your data and ---Try Again---'".format(ip_address))
            exit(-1)
    return ip_address


# ***********************************************


# user defined function:
# name: login_to_host(tn, string_tag)
# input: tn = A telnet object
#        string_tag =  Information associated to the current host
# function: returns the telnet session object after correct login
def login_to_host(tn, string_tag):
    global eColumn
    while (True):
        # host_ip = ip_input_format("Setup for 42XX / 9XX node {}\n Enter the host ip: ".format(count))
        print(string_tag)
        host_ip = ip_input_format(sheetN.cell(row=eRow, column=eColumn).value)
        print(" Enter the host ip: {}".format(host_ip))

        telnet_host_flag = 0
        try:
            tn = telnetlib.Telnet(host_ip)
        except:
            print("Host could not be found. Update/Correct your data and ---Try Again----")
            telnet_host_flag = 1
            exit(-1)
            # tn.close()
            continue
        break
    print(tn)

    eColumn = eColumn + 1
    user_name = sheetN.cell(row=eRow, column=eColumn).value
    eColumn = eColumn + 1
    password = sheetN.cell(row=eRow, column=eColumn).value
    last_line = tn.read_until(b'Username: ', timeout=1)
    last_line = last_line.splitlines()[-1]

    while (True):
        # last_line = tn.read_until(b'Username: ')
        # last_line = last_line.splitlines()[-1]
        flag = 0
        # print(last_line)
        if "Username" in last_line.decode('ascii'):
            print(" Enter username: " + user_name)
            tn.write((user_name + '\n').encode('ascii'))
            flag = 1

        if (flag == 1):
            last_line = tn.read_until(b'Password: ', timeout=1)
            last_line = last_line.splitlines()[-1]

        if "Password" in last_line.decode('ascii'):
            print(" Enter password: " + password)
            tn.write((password + "\n").encode('ascii'))

        last_line = tn.read_until(b'#', timeout=10)  # change this timeout if required
        # print(last_line)
        last_line = last_line.splitlines()[-1]
        # print(last_line)

        if "#" in last_line.decode('ascii') or ">" in last_line.decode('ascii'):
            break
        # else:
        #   continue
        print("\r------LOGIN ERROR--TRY AGAIN------\n")

    eColumn = eColumn + 1
    enablePassword = sheetN.cell(row=eRow, column=eColumn).value
    if ">" in last_line.decode('ascii'):
        tn.write(("en\n").encode('ascii'))
        while (True):
            tn.read_until(b'Password: ', timeout=1)
            print(" Enable password: " + enablePassword)
            tn.write((enablePassword + "\n").encode('ascii'))
            last_line = tn.read_until(b'#', timeout=1)
            if "#" in last_line.decode('ascii'):
                break
            print(" Error in enable password, update/correct the data, TRY again")

    return tn


# *******************************************

# user defined function:
# name: obtain_network_info(mpls_label_range_info, rsvp_bw_input_info)
# input: IOS_TPYE = either IOS_XR or IOS_XE
#        mpls_label_range_info = Just to show the valid mpls label range for input
#        rsvp_bw_input_info = rsvp bandwidth input info
# function: returns network information given below (per node):
#            1. Loopback IP
#            2. OSPF process ID
#            3. Number of interfaces to setup
#            4. Interface list
#            5. IP address list
#            6. Subnet mask list
#            7. Check value for using default/user-defined mpls label range
#            8. minimum label value if check value in 7 refers to user-defined
#            9. maximum label value of check value in 7 refers to user-defined
def obtain_network_info(IOS_TYPE, mpls_label_range_info, rsvp_bw_input_info):
    global eColumn
    eColumn = eColumn + 1
    LoopbackIp = ip_input_format(sheetN.cell(row=eRow, column=eColumn).value)
    print(" Enter Loopback ip: " + LoopbackIp)

    eColumn = eColumn + 1
    No_of_interfaces = sheetN.cell(row=eRow, column=eColumn).value
    print(" Enter the number of Interfaces for setup: {}".format(No_of_interfaces))

    eColumn = eColumn + 1
    int_list = sheetN.cell(row=eRow, column=eColumn).value.split()
    # int_list.insert(0, 'Interface Name')

    eColumn = eColumn + 1
    ip_list = sheetN.cell(row=eRow, column=eColumn).value.split()
    # ip_list.insert(0, 'IP Address')

    eColumn = eColumn + 1  # This section
    if IOS_TYPE == IOS_XE:
        subnet_value = ip_input_format(sheetN.cell(row=eRow, column=eColumn).value)  # may change based on
    else:
        subnet_value = sheetN.cell(row=eRow, column=eColumn).value

    subnet_list = list()  # the input from excel file
    for i in range(1, No_of_interfaces+1):
        subnet_list.append(subnet_value)

    # subnet_list = sheetN.cell(row=eRow, column=eColumn).value.split()
    # subnet_list.insert(0, 'Subnet Mask')

    t = Texttable(max_width=150)
    t.add_row(int_list)
    t.add_row(ip_list)
    t.add_row(subnet_list)

    print(t.draw())

    eColumn = eColumn + 1
    router_ospf_process_id = sheetN.cell(row=eRow, column=eColumn).value
    print(" Enter Router OSPF process ID: {}".format(router_ospf_process_id))

    eColumn = eColumn + 1
    defaultMplsLabelRangeCheck = sheetN.cell(row=eRow, column=eColumn).value
    print(" Do you want to keep default mpls label[Y/N]? [Y]: " + defaultMplsLabelRangeCheck)

    eColumn = eColumn + 1
    labelRange = sheetN.cell(row=eRow, column=eColumn).value.split()

    if (defaultMplsLabelRangeCheck.upper() == "N"):
        minLabel = labelRange[0]
        maxLabel = labelRange[2]
        # print("\tMPLS label range<16-32768>: {} - {}".format(minLabel, maxLabel))
        print("\t" + mpls_label_range_info + ": {} - {}".format(minLabel, maxLabel))
    else:
        minLabel = 0
        maxLabel = 0

    eColumn = eColumn + 1
    rsvp_input_bw_percent = str(sheetN.cell(row=eRow, column=eColumn).value).split('\n')
    bandLen = len(rsvp_input_bw_percent)
    if (bandLen==1):
        temp = rsvp_input_bw_percent[0]
        for indx in range(1, No_of_interfaces):
            rsvp_input_bw_percent.append(temp)
    elif (bandLen == No_of_interfaces):
        rsvp_input_bw_percent = rsvp_input_bw_percent
    else:
        print("Rsvp bandwidth Entry is invalid/missing. Update you data and Try Again")
        exit(-1)

    setRSVPBandWidthString = list()
    for indx in range(1, No_of_interfaces+1):
        if "%" in rsvp_input_bw_percent[indx-1]:
            setRSVPBandWidthString.append('percent {}'.format(rsvp_input_bw_percent[indx-1])[:-1])
        else:
            setRSVPBandWidthString.append('{}'.format(rsvp_input_bw_percent[indx-1]))
        print(int_list[indx-1] + ': ' + rsvp_bw_input_info + rsvp_input_bw_percent[indx-1])

    # print(rsvp_bw_input_info + rsvp_input_bw_percent)
    return LoopbackIp, router_ospf_process_id, No_of_interfaces, \
           int_list, ip_list, subnet_list, defaultMplsLabelRangeCheck, \
           setRSVPBandWidthString, minLabel, maxLabel


# *************************************

# ............... Starting of the main code .....................
No_of_4200_Nodes = sheetN.cell(row=eRow, column=eColumn).value
print("Number of NCS 42XX / ASR 9XX in the Topology: {}".format(No_of_4200_Nodes))

#####*****Starting with NCS 4200/ ASR 9XX Setup*****#######
if (No_of_4200_Nodes <= 0):
    print("No NCS 42XXs/ ASR 9XXs for Setup\n")
else:
    print("*********Starting 42XX/ASR 9XXs Setup***********")

    for count in range(1, No_of_4200_Nodes + 1):
        tn = telnetlib.theNULL
        eRow = eRow + 1
        eColumn = 3

        tn = login_to_host(tn, "Setup for 42XX / 9XX node {}".format(count))
        MPLS_label_range_info = "MPLS label range<16-32768>"
        RSVP_bandwidth_input_info = " Enter rsvp bandwidth(kbps) in Number or Percentage(use % with input):"
        # network_info = obtain_network_info(MPLS_label_range_info, RSVP_bandwidth_input_info)
        # print(network_info)

        LoopbackIp, router_ospf_process_id, No_of_interfaces, int_list, \
        ip_list, subnet_list, defaultMplsLabelRangeCheck, setRSVPBandWidthString, \
        minLabel, maxLabel = obtain_network_info(IOS_XE, MPLS_label_range_info, RSVP_bandwidth_input_info)

        eColumn = eColumn + 1
        check_external_clock_source = sheetN.cell(row=eRow, column=eColumn).value
        print(" Does this node have external clock source? [Y/N]? [N]: " + check_external_clock_source)

        clock_setup_commands = ''

        if (check_external_clock_source.upper() == "Y"):
            eColumn = eColumn + 1
            no_of_bitsSource = sheetN.cell(row=eRow, column=eColumn).value
            print("\tNumber of bits source: {}".format(no_of_bitsSource))
            eColumn = eColumn + 1
            priority = sheetN.cell(row=eRow, column=eColumn).value
            eColumn = eColumn + 1
            RO = sheetN.cell(row=eRow, column=eColumn).value
            print("\tEnter the external clock-source: " + RO)
            clock_setup_commands = 'network-clock input-source {} external {}\n'.format(priority, RO)
        else:
            eColumn = eColumn + 1
            no_of_interfaces_for_clock_source = sheetN.cell(row=eRow, column=eColumn).value
            print(
                " Enter the number of interfaces for input clock source: {}".format(no_of_interfaces_for_clock_source))
            clock_setup_commands = 'no network-clock input-source\nnetwork-clock revertive\n'
            eColumn = eColumn + 1
            var1 = str(sheetN.cell(row=eRow, column=eColumn).value).split()
            eColumn = eColumn + 1
            var2 = sheetN.cell(row=eRow, column=eColumn).value.split()
            for i in range(1, no_of_interfaces_for_clock_source + 1):
                print("\tEnter priority 1st, then space, and interface name for interface {}: {} {} ".format(i, var1[i - 1],
                                                                                                          var2[i - 1]))
                clock_setup_commands = clock_setup_commands + 'network-clock input-source {} interface {}\n'.format(
                    var1[i - 1], var2[i - 1])

        # print(clock_setup_commands)

        # tn.write(("en\n").encode('ascii'))
        # tn.read_until(b'Password: ')
        # tn.write((password + "\n").encode('ascii'))

        #####****************Commands start here****************#######

        if (defaultMplsLabelRangeCheck.upper() == "N"):
            tn.write(("conf t\n"
                      "mpls label range %s %s\n" % (minLabel, maxLabel) +
                      "mpls ldp label\n"
                      "allocate global host-routes\n"
                      "end\n").encode('ascii'))

        #Commands covering global level and loopback level
        tn.write(("\n"
                  "conf t\n"
                  "cdp run\n"
                  "esmc process\n"
                  "network-clock synchronization automatic\n"
                  "network-clock synchronization ssm option 2 GEN2\n"
                  "network-clock synchronization mode QL-enabled\n"
                  "int loopback0\n"
                  "ip address %s 255.255.255.255\n" %LoopbackIp +
                  "exit\n"
                  "mpls ldp router-id loopback0 force\n"
                  "mpls traffic-eng tunnels\n"
                  "Router ospf %s\n" %router_ospf_process_id +
                  "router-id %s\n" % LoopbackIp +
                  "passive-interface Loopback0\n"
                  "mpls traffic-eng area 0\n"
                  "mpls traffic-eng router-id loopback0\n"
                  "int Loopback0\n"
                  "ip ospf %s area 0\n" %router_ospf_process_id +
                  "end\n").encode('ascii'))

        ##Commands covering interface level
        for ints in range(1,  No_of_interfaces+1):
            tn.write(("conf t\n"
                     "int %s\n" %int_list[ints-1] +
                     "cdp enable\n"
                     "ip address %s %s\n" %(ip_list[ints-1], subnet_list[ints-1]) +
                     "ip ospf %s area 0\n" %router_ospf_process_id +
                     "ip ospf network point-to-point\n"
                     "no ospf network point-to-point\n"
                     "mpls ip\n"
                     "mpls traffic-eng tunnels\n" +
                     "ip rsvp bandwidth "+ setRSVPBandWidthString[ints-1] + "\n"+
                     "synchronous mode\n"
                     "logging event link-status\n"
					 "no shutdown\n"
                     "end\n").encode('ascii'))

        ##Commands for network clock syncronization
        tn.write(("conf t\n" +
                  clock_setup_commands + "\n" +
                  "end\n").encode('ascii'))

        tn.write(("wr\n").encode('ascii'))
        tn.write(("exit\n").encode('ascii'))
        #####****************Commands ends here****************#######
        print(tn.read_all().decode('ascii'))
        tn.close()


indent = 4
print(
    __doc__,
    "Author:",
    " " * indent + __author__,
    "",
    "Collaborator:",
    " " * indent + __collaborator__,
    "",
    __copyright__,
    "Licensed Under: " + __license__,
    sep="\n"
)
